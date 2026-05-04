from __future__ import annotations

import json
import unicodedata
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "insumos" / "documentacion_sistema_sgs_iso21101.xlsx"
OUTPUT = ROOT / "app" / "form_catalog.js"

FORM_REQUIREMENT_MAP = {
    "organizaciones": "4.1",
    "usuarios": "5.3",
    "capitulos_iso": "4.4",
    "compromiso_liderazgo": "5.1",
    "contexto_interno_externo": "4.1",
    "contexto_actividades": "4.3",
    "impacto_riesgo": "6.1.2",
    "probabilidad_riesgo": "6.1.2",
    "mapa_riesgos": "6.1.2",
    "plantillas_documentos": "7.5",
    "procedimientos_sgs": "4.4",
    "politica_seguridad": "5.2",
    "roles_responsabilidades": "5.3",
    "normograma": "6.1.3",
    "politica_sgs": "5.2",
    "objetivos_sgs": "6.2",
    "plan_objetivos_sgs": "6.2",
    "cargos": "5.3",
    "competencias_minimas": "7.2",
    "personas": "7.2",
    "compromiso_personal": "7.3",
    "plan_comunicaciones": "7.4",
    "registro_comunicaciones_seguridad": "7.4",
    "documentos_controlados": "7.5",
    "documentacion_minima": "7.5",
    "politica_datos_personales": "7.4",
    "bitacora_actividades": "8.1",
    "procedimientos_servicio": "8.1",
    "equipos": "7.1",
    "plan_emergencia": "8.2",
    "simulacros": "8.2",
    "simulacro_evidencias": "8.2",
    "registro_participantes_evidencia": "7.4",
    "registro_incidentes": "8.3",
    "acciones_sugeridas": "10.2",
    "seguimiento_medicion": "9.1",
    "programa_auditorias": "9.2",
    "plan_auditoria": "9.2",
    "informe_auditoria_detalle": "9.2",
    "referencia_norma_iso21101": "4.4",
    "acciones_correctivas": "10.1",
    "informe_auditoria": "9.2",
    "seguimiento_acciones_correctivas": "10.1",
    "revision_direccion": "9.3",
    "procedimientos_obligatorios": "4.4",
    "procedimientos_implementados": "8.1",
}


def clean(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return " ".join(text.split())


def ascii_text(value: str) -> str:
    normalized = unicodedata.normalize("NFKD", value)
    return normalized.encode("ascii", "ignore").decode("ascii")


def main() -> None:
    workbook = load_workbook(WORKBOOK, data_only=True)
    tables_sheet = workbook["Tablas"]
    fields_sheet = workbook["Campos"]

    forms: OrderedDict[str, dict] = OrderedDict()
    for row in tables_sheet.iter_rows(min_row=2, values_only=True):
        table = clean(row[0])
        if not table:
            continue
        key = ascii_text(table).strip()
        if key not in forms:
            forms[key] = {
                "table": key,
                "title": key.replace("_", " ").title(),
                "requirement": FORM_REQUIREMENT_MAP.get(key, "4.4"),
                "description": ascii_text(clean(row[1])),
                "operation": ascii_text(clean(row[2] if len(row) > 2 else "")),
                "notes": ascii_text(clean(row[3] if len(row) > 3 else "")),
                "fields": [],
            }
        else:
            extra_note = ascii_text(clean(row[3] if len(row) > 3 else ""))
            if extra_note and extra_note not in forms[key]["notes"]:
                forms[key]["notes"] = (forms[key]["notes"] + " " + extra_note).strip()

    for row in fields_sheet.iter_rows(min_row=2, values_only=True):
        table = ascii_text(clean(row[0]))
        field = ascii_text(clean(row[1]))
        if not table or not field or field.lower() == "campo":
            continue
        if table not in forms:
            forms[table] = {
                "table": table,
                "title": table.replace("_", " ").title(),
                "requirement": FORM_REQUIREMENT_MAP.get(table, "4.4"),
                "description": "",
                "operation": "",
                "notes": "",
                "fields": [],
            }
        field_names = {item["name"] for item in forms[table]["fields"]}
        if field in field_names:
            continue
        forms[table]["fields"].append(
            {
                "name": field,
                "label": field.replace("_", " ").title(),
                "description": ascii_text(clean(row[2])),
                "type": ascii_text(clean(row[3] if len(row) > 3 else "")) or "TEXT",
            }
        )

    catalog = list(forms.values())
    payload = json.dumps(catalog, ensure_ascii=True, indent=2)
    OUTPUT.write_text(f"window.formCatalog = {payload};\n", encoding="utf-8")
    print(f"Wrote {len(catalog)} forms and {sum(len(item['fields']) for item in catalog)} fields to {OUTPUT}")


if __name__ == "__main__":
    main()
